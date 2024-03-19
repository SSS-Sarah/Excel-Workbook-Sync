# This python script assumes that a localCompare excel workbook already exists.
# Press Ctrl + 0 to stop this script

import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEvent, FileSystemEventHandler
import sys
import keyboard
import contextlib
import tkinter as tk 

window = None
message_label = None

# Calls update_local_compare() func if master is modified
class FileChangeHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path.endswith('MasterCompare.xlsx'):
            update_local_compare()
            #print("Master changed; updated local.") 


def update_local_compare():
    # Assumes both workbooks are in the current directory
    master_compare_workbook  = openpyxl.load_workbook("MasterCompare.xlsx") # replace with filename
    local_compare_workbook = openpyxl.load_workbook("LocalCompare.xlsx") # replace with filename

    # Common sheets: CCA, SCA, TCA
    common_sheets = ['CCA', 'SCA', 'TCA']

    # For each common sheet, update local as master is updated
    # note: only columns are added in master, not rows
    for sheet_name in common_sheets:
        master_sheet = master_compare_workbook[sheet_name]
        local_sheet = local_compare_workbook[sheet_name]
        
        # Find new col in master, add them to same position in local
        for column_index, column in enumerate(master_sheet.iter_cols(values_only=True), start=1):
            local_column = local_sheet.cell(row=1, column=column_index).column_letter
            
            for cell in column:
                local_sheet[f"{local_column}{cell.row}"] = cell.value


    # note: add support for changes in rows in the master LATER/EXTRA
    local_compare_workbook.save('LocalCompare.xlsx')


def handle_interrupt(frame):
    
    # Display in GUI
    global window
    message_label.config(text="Interrupt signal received. Script is ending...")
    window.quit()
    # print("Interrupt signal received. Script is ending...")
    sys.exit(0)
    

# Handle interruption and script runtime
if __name__ == "__main__":
    
    # For script exit message
    window = tk.Tk()
    message_label = tk.Label(window, text="")
    message_label.pack()
    
    custom_interrupt_key = "ctrl + 0"  # custom interruption key combo
    window.mainloop() # Start tkinter main loop
    keyboard.add_hotkey(custom_interrupt_key, handle_interrupt)

    event_handler = FileChangeHandler()
    observer = Observer()
    observer.schedule(event_handler, path='.', recursive=False)
    observer.start()

    # Keeps running script until interrupted.
    with contextlib.suppress(SystemExit):
        while True:
            pass
    observer.stop()
    observer.join()
